from django.utils import timezone
import datetime as DT
from django.db.models import Q
from rest_framework.exceptions import NotAcceptable
import openpyxl

from apps.billing.config import PACKAGE_INFO
from apps.user.config import ROLE_DICT
from apps.user.models import User
from apps.state.models import UserState
from apps.task.models import Task
from apps.location.models import ClientLocation
from .graph_check import Graph

from apps.user.auth_helpers import create_username, check_username_exists


def clean_whitespace(cell):
    value = str(cell.value) if cell.value else ''
    if len(value) > 0:
        value = ''.join(value.split())
    return value


def clean_whs(value):
    return ''.join(value.split())


def strip_whitespace(cell):
    value = str(cell.value) if cell.value else ''
    if len(value) > 0:
        value = value.strip()
    return value


def validate_excel(excel_file):
    if not excel_file:
        msg = 'File is not attached'
        raise NotAcceptable(detail=msg)
    if not excel_file.name.endswith('.xlsx'):
        msg = 'File is not excel type'
        raise NotAcceptable(detail=msg)

    if excel_file.multiple_chunks():
        msg = "Uploaded file is too big (%.2f MB)." % (excel_file.size / (1000 * 1000),)
        raise NotAcceptable(detail=msg)


def validate_csv(csv_file):
    if not csv_file:
        msg = 'File is not attached'
        raise NotAcceptable(detail=msg)
    if not csv_file.name.endswith('.csv'):
        msg = 'File is not csv type'
        raise NotAcceptable(detail=msg)

    if csv_file.multiple_chunks():
        msg = "Uploaded file is too big (%.2f MB)." % (csv_file.size / (1000 * 1000),)
        raise NotAcceptable(detail=msg)


def check_csv_fields(csv_file, input_fields, delimiter):
    file_data = csv_file.read().decode('utf-8')
    csv_row_list = []
    lines = file_data.split("\n")
    csv_fields = ["".join(col.split()) for col in lines[0].split(delimiter)]

    for idx in range(1, len(lines)):
        value_list = lines[idx].split(delimiter)
        if len(value_list) == len(csv_fields):
            cur_dict = dict(zip(csv_fields, value_list))
            csv_row_list.append(cur_dict)

    csv_field_set = set(csv_fields)
    input_field_set = set(input_fields)
    csv_diff = csv_field_set.difference(input_field_set)
    if csv_field_set == set() or csv_diff != set():
        return False, None
    return csv_diff == set(), csv_row_list


def invalid_office_time(validated_data):
    start = validated_data['day_start']
    end = validated_data['day_end']
    diff = DT.datetime.combine(DT.date.today(), end) \
        - DT.datetime.combine(DT.date.today(), start)
    if diff < timezone.timedelta(hours=1):
        return True
    return False


def get_org_details(org):
    package = org.subscription.current_usage.package
    renew_needed = org.subscription.renew_needed
    return {
        'packages': PACKAGE_INFO,
        'oid': org.oid,
        'org_id': org.id,
        'org_name': org.org_name,
        'org_address': org.address,
        'package_info': PACKAGE_INFO[package] if (package and package in [1, 2, 3]) else None,

        'renew_needed': renew_needed,
        'day_start': str(org.day_start),
        'day_end': str(org.day_end),
        'location_interval': org.location_interval,
        'min_work_hour_p': org.min_work_hour_p,
        # 'min_task_hour_p': org.min_task_hour_p,
        'weekend': org.weekend,
        'logo': org.logo,
        'tracking_enabled': org.tracking_enabled,
    }


def set_duration(hours):
    return timezone.now() + timezone.timedelta(hours=hours)


def get_common_agents_org(organizer):
    return Q(org=organizer.org) & Q(role=ROLE_DICT['Employee']) & Q(is_active=True)


def get_location_dict(name, address, lat, lng):
    return {
        'name': name,
        'address': address,
        'point': {
            "lat": float(lat),
            "lon": float(lng)
        }
    }


def get_location_obj(loc):
    return {
        'id': loc.id,
        'name': loc.name,
        'address': loc.address,
        'point': loc.point
    }


def get_location_list(locs):
    location_list = []
    for loc in locs:
        loc_details = get_location_obj(loc)
        location_list.append(loc_details)
    return location_list


def check_team_hierarchy_excel(excel_file, organizer):
    oid = organizer.org.oid
    roles = ['Manager', 'Employee']
    user_graph = Graph()

    user_list = []
    new_users = []
    worksheet = openpyxl.load_workbook(excel_file)['Sheet1']
    for row in worksheet.iter_rows(min_row=2):
        if row is None:
            break
        # print(row)
        user_handle = clean_whitespace(row[0])
        if len(user_handle) < 1:
            err_msg = 'Empty employee id!'
            raise NotAcceptable(detail=err_msg)

        if not user_handle.isalnum():
            err_msg = 'Invalid employee id: <' + user_handle + '>'
            raise NotAcceptable(detail=err_msg)

        name = row[1].value
        phone = row[2].value
        email = row[3].value
        designation = row[4].value
        role = row[5].value
        if role not in roles:
            err_msg = 'Invalid user role: ' + role
            raise NotAcceptable(detail=err_msg)

        manager = clean_whitespace(row[6])
        if len(manager) > 0:
            if not manager.isalnum():
                err_msg = 'Invalid manager id: <' + manager + '>'
                raise NotAcceptable(detail=err_msg)
            user_graph.addEdge(manager, user_handle)

        full_username = create_username(oid, user_handle)
        user_qs = User.objects.filter(Q(username=full_username))
        if not user_qs.exists():
            new_users.append(full_username)

        user_list.append({
            'username': user_handle,
            'full_name': name,
            'phone': phone,
            'email': email,
            'designation': designation,
            'role': ROLE_DICT[role],
            'parent': manager,
        })

    if user_graph.isCyclic():
        # print(user_graph.graph)
        err_msg = 'Invalid user hierarchy!'
        raise NotAcceptable(detail=err_msg)

    added = organizer.org.subscription.added_agents
    new_agent = len(new_users)
    limit = organizer.org.subscription.agent_limit
    if new_agent + added > limit:
        diff = (new_agent + added) - limit
        err_msg = 'Employee limit exceeded! Please purchase another ' + str(diff) + ' accounts!'
        raise NotAcceptable(detail=err_msg)
    # print(user_list)
    if len(user_list) < 1:
        err_msg = 'Invalid excel file!'
        raise NotAcceptable(detail=err_msg)
    return user_list


def check_team_hierarchy_csv(csv_file, organizer):
    oid = organizer.org.oid
    roles = ['Manager', 'Employee']
    user_graph = Graph()

    user_list = []
    new_users = []
    input_fields = ['employee_id', 'name', 'phone', 'email', 'designation', 'role', 'manager_id']
    flag, csv_dict = check_csv_fields(csv_file, input_fields, ';')
    if not flag:
        err_msg = 'Invalid csv file!'
        raise NotAcceptable(detail=err_msg)

    for row in csv_dict:

        user_handle = clean_whs(row['employee_id'])
        if len(user_handle) < 1:
            err_msg = 'Empty employee id!'
            raise NotAcceptable(detail=err_msg)

        if not user_handle.isalnum():
            err_msg = 'Invalid employee id: <' + user_handle + '>'
            raise NotAcceptable(detail=err_msg)

        name = row['name']
        phone = row['phone']
        email = row['email']
        designation = row['designation']
        role = row['role']
        if role not in roles:
            err_msg = 'Invalid user role: ' + role
            raise NotAcceptable(detail=err_msg)

        manager = clean_whs(row['manager_id'])
        if len(manager) > 0:
            if not manager.isalnum():
                err_msg = 'Invalid manager id: <' + manager + '>'
                raise NotAcceptable(detail=err_msg)
            user_graph.addEdge(manager, user_handle)

        full_username = create_username(oid, user_handle)
        user_qs = User.objects.filter(Q(username=full_username))
        if not user_qs.exists():
            new_users.append(full_username)

        user_list.append({
            'username': user_handle,
            'full_name': name,
            'phone': phone,
            'email': email,
            'designation': designation,
            'role': ROLE_DICT[role],
            'parent': manager,
        })

    if user_graph.isCyclic():
        # print(user_graph.graph)
        err_msg = 'Invalid user hierarchy!'
        raise NotAcceptable(detail=err_msg)

    added = organizer.org.subscription.added_agents
    new_agent = len(new_users)
    limit = organizer.org.subscription.agent_limit
    if new_agent + added > limit:
        diff = (new_agent + added) - limit
        err_msg = 'Employee limit exceeded! Please purchase another ' + str(diff) + ' accounts!'
        raise NotAcceptable(detail=err_msg)
    # print(user_list)
    if len(user_list) < 1:
        err_msg = 'Invalid excel file!'
        raise NotAcceptable(detail=err_msg)
    return user_list


def check_tasks_excel(excel_file, user):
    oid = user.org.oid
    task_list = []
    worksheet = openpyxl.load_workbook(excel_file)['Sheet1']
    for row in worksheet.iter_rows(min_row=2):
        if row is None:
            break
        try:
            title = strip_whitespace(row[0])
            duration = int(row[1].value)
            address = strip_whitespace(row[2])
            manager_id = clean_whitespace(row[3])

            if len(manager_id) < 1:
                err_msg = 'Manager not provided!'
                raise NotAcceptable(detail=err_msg)

            manager_qs = User.objects.filter(
                Q(username=create_username(oid, manager_id))
            )
            if not manager_qs.exists():
                err_msg = 'Manager: ' + manager_id + ' not found!'
                raise NotAcceptable(detail=err_msg)
            manager_obj = manager_qs[0]

            agents_str = clean_whitespace(row[4]).split(',')
            agent_usernames = list(filter(None, agents_str))
            agent_objs = []

            for username in agent_usernames:
                agent_qs = User.objects.filter(
                    Q(username=create_username(oid, username))
                )
                if not agent_qs.exists():
                    err_msg = 'Agent: ' + username + ' not found!'
                    raise NotAcceptable(detail=err_msg)

                agent = agent_qs[0]
                if agent.parent != manager_obj:
                    err_msg = 'Agent: ' + username + ' is not related to manager: ' + manager_id + '!'
                    raise NotAcceptable(detail=err_msg)
                agent_objs.append(agent_qs[0])

            image_req = bool(int(row[5].value))
            demand = bool(int(row[6].value))
            other_fields = str(row[7].value).split(';')

        except Exception as e:
            err_msg = str(e)
            raise NotAcceptable(detail=err_msg)

        task_list.append({
            'title': title,
            'duration': duration,
            'address': address,
            'manager': manager_obj,
            'agents': agent_objs,
            'image_req': image_req,
            'demand': demand,
            'other_fields': other_fields,
        })

    # print(task_list)
    new_task = len(task_list)
    task_limit = user.org.subscription.task_limit
    consumed_tasks = user.org.subscription.current_usage.consumed_tasks
    if new_task + consumed_tasks > task_limit:
        diff = (new_task + consumed_tasks) - task_limit
        err_msg = 'Task limit exceeded! Please upgrade for another ' + str(diff) + ' tasks!'
        raise NotAcceptable(detail=err_msg)

    return task_list


def check_tasks_csv(csv_file, organizer):
    oid = organizer.org.oid
    task_list = []
    input_fields = ['title', 'duration', 'address',	'manager_id', 'image_required', 'collect_demand', 'other_fields']
    flag, csv_dict = check_csv_fields(csv_file, input_fields, ';')
    for row in csv_dict:
        if row is None:
            break
        try:
            title = row['title']
            duration = int(row['duration'])
            address = row['address']
            manager = clean_whs(row['manager_id'])

            if len(manager) < 1:
                err_msg = 'Manager not provided!'
                raise NotAcceptable(detail=err_msg)

            manager_qs = User.objects.filter(
                Q(username=create_username(oid, manager))
            )
            if not manager_qs.exists():
                err_msg = 'Manager: ' + manager + ' not found!'
                raise NotAcceptable(detail=err_msg)
            manager = manager_qs[0]

            image_req = bool(int(row['image_required']))
            demand = bool(int(row['collect_demand']))
            other_fields = str(row['other_fields']).split(',')

        except Exception as e:
            err_msg = str(e)
            raise NotAcceptable(detail=err_msg)

        task_list.append({
            'title': title,
            'duration': duration,
            'address': address,
            'manager': manager,
            'agents': [],
            'image_req': image_req,
            'demand': demand,
            'other_fields': other_fields,
        })

    # print(task_list)
    new_task = len(task_list)
    task_limit = organizer.org.subscription.task_limit
    consumed_tasks = organizer.org.subscription.current_usage.consumed_tasks
    if new_task + consumed_tasks > task_limit:
        diff = (new_task + consumed_tasks) - task_limit
        err_msg = 'Task limit exceeded! Please upgrade for another ' + str(diff) + ' tasks!'
        raise NotAcceptable(detail=err_msg)

    return task_list


def check_locations_excel(excel_file):
    location_list = []
    worksheet = openpyxl.load_workbook(excel_file)['Sheet1']
    for row in worksheet.iter_rows(min_row=2):
        if row is None:
            break
        # print(row)
        name = strip_whitespace(row[0])
        address = strip_whitespace(row[1])
        lat = clean_whitespace(row[2])
        lon = clean_whitespace(row[3])

        if len(name) < 1 or len(lat) < 1 or len(lon) < 1:
            err_msg = 'Name, lat and lon values cannot be empty!'
            raise NotAcceptable(detail=err_msg)

        location_list.append(get_location_dict(name, address, lat, lon))

    # print(location_list)
    return location_list


def check_locations_csv(csv_file):
    location_list = []
    input_fields = ['name', 'address', 'latitude', 'longitude']
    flag, csv_dict = check_csv_fields(csv_file, input_fields, ';')
    for row in csv_dict:

        name = row['name'].strip()
        address = row['address']
        lat = clean_whs(row['latitude'])
        lon = clean_whs(row['longitude'])

        if len(name) < 1 or len(lat) < 1 or len(lon) < 1:
            err_msg = 'Name, latitude and longitude values cannot be empty!'
            raise NotAcceptable(detail=err_msg)

        location_list.append(get_location_dict(name, address, lat, lon))

    # print(location_list)
    return location_list


def create_teams_sync(user_list, organizer):
    oid = organizer.org.oid
    new_agents = 0

    for user_dict in user_list:
        cur_dict = user_dict.copy()
        cur_dict.pop('parent')
        if not check_username_exists(oid, cur_dict['username']):
            cur_dict['username'] = create_username(oid, cur_dict['username'])
            user = User(**cur_dict)
            user.org = organizer.org
            user.is_active = True
            user.save()
            userstate = UserState.objects.create(
                user=user
            )
            new_agents += 1
        else:
            username = create_username(oid, cur_dict['username'])
            cur_dict.pop('username')
            User.objects.filter(
                Q(username=username)
            ).update(**cur_dict)

    for user_dict in user_list:
        # print(user_dict)
        if len(user_dict['parent']) > 0:
            manager_qs = User.objects.filter(
                username=create_username(oid, user_dict['parent'])
            )
            if manager_qs.exists():
                parent = manager_qs[0]
                user = User.objects.get(
                    username=create_username(oid, user_dict['username'])
                )
                try:
                    user.parent = parent
                    user.save()
                except Exception as e:
                    raise NotAcceptable(detail=str(e))
            else:
                err_msg = 'Manager : ' + user_dict['parent'] + ' not found!'
                raise NotAcceptable(detail=err_msg)
    organizer.org.subscription.added_agents += new_agents
    organizer.org.subscription.save()

    return True


def create_locations_sync(location_list, organizer):
    for loc in location_list:
        new_loc = ClientLocation(
            name=loc['name'],
            address=loc['address'],
            point=loc['point'],
            org=organizer.org
        )
        new_loc.save()


def create_tasks_sync(task_list, organizer):
    for task in task_list:
        # print(task['agents'])
        agents = task['agents']

        new_task = Task(
            manager=task['manager'],
            title=task['title'],
            address=task['address'],
            deadline=set_duration(task['duration'])
        )
        new_task.task_type = 'Custom Upload'
        custom_fields = {}
        if task['image_req']:
            new_task.image_required = True
        if task['demand']:
            custom_fields.update({
                'demand_keys': ['Item', 'Quantity'],
                'Demands': []
            })
        for field in task['other_fields']:
            custom_fields.update({
                field: ''
            })
        new_task.custom_fields = custom_fields
        new_task.save()
        if len(agents) > 1:
            new_task.agent_list.add(*agents)
            new_task.save()

    organizer.org.subscription.current_usage.consumed_tasks += len(task_list)
    organizer.org.subscription.current_usage.save()
